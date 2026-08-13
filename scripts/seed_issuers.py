"""Carga inicial de emissores: taxonomia + ratings.

Fase 1 do Hub Credit Research (04/08/2026). Lê a planilha
`Taxonomia_Emissores.xlsx` (gerada a partir do Dashboard_Snapshot e
revisada pelo Allan) e, opcionalmente, o histórico de ações de rating.

Uso:

    # simulação -- não grava nada, só mostra o que faria
    python -m scripts.seed_issuers --taxonomia "../Taxonomia_Emissores.xlsx" --dry-run

    # carga de verdade
    python -m scripts.seed_issuers --taxonomia "../Taxonomia_Emissores.xlsx"

    # + histórico de ratings do checkpoint do scraper
    python -m scripts.seed_issuers --taxonomia "..." --ratings data/mapeamento_ratings_checkpoint.jsonl

    # só recalcular o rating vigente (sem carregar nada)
    python -m scripts.seed_issuers --apenas-recalcular

IDEMPOTENTE: pode rodar quantas vezes quiser. Emissor já existente é
atualizado (respeitando precedência de origem — edição MANUAL na
Administração nunca é sobrescrita por uma recarga), ação de rating já
gravada é ignorada.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.db import Base, SessionLocal, engine, run_migrations  # noqa: E402
from app.models import Issuer  # noqa: E402
from app.spreads import issuers as isv  # noqa: E402
from app.spreads.issuer_key import issuer_key  # noqa: E402

# Nome das colunas na aba "Emissores" da planilha. Se o Allan renomear,
# ajustar aqui (de propósito não é "primeira coluna, segunda coluna" —
# posição quebra em silêncio quando alguém insere uma coluna no meio).
COL_NOME = "Nome Anbima"
COL_SETOR = "Setor"
COL_SUB = "Subsetor"
COL_GRUPO = "Grupo Econômico"

AGENCIA_POR_CAMPO = {"fitch": "FITCH", "sp": "SP", "moodys": "MOODYS"}


def _texto(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ("", "-", "N.A.", "nan", "None") else s


def _data(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Taxonomia (planilha)
# ---------------------------------------------------------------------------

def carregar_taxonomia(db, caminho: Path, *, dry_run: bool) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    if "Emissores" not in wb.sheetnames:
        raise SystemExit(
            f"{caminho.name}: aba 'Emissores' não encontrada (abas: {wb.sheetnames})"
        )
    ws = wb["Emissores"]

    linhas = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(linhas)]
    faltando = [c for c in (COL_NOME, COL_SETOR, COL_SUB, COL_GRUPO) if c not in header]
    if faltando:
        raise SystemExit(f"{caminho.name}: colunas ausentes na aba Emissores: {faltando}")
    idx = {c: header.index(c) for c in (COL_NOME, COL_SETOR, COL_SUB, COL_GRUPO)}

    criados = atualizados = sem_taxonomia = ignorados = 0
    chaves_vistas: set[str] = set()
    colisoes: list[tuple[str, str]] = []

    for row in linhas:
        nome = _texto(row[idx[COL_NOME]])
        if not nome:
            ignorados += 1
            continue
        chave = issuer_key(nome)
        if not chave:
            ignorados += 1
            continue
        if chave in chaves_vistas:
            # Duas linhas da planilha colapsam na mesma chave. Não é erro
            # fatal (costuma ser a mesma empresa escrita de dois jeitos),
            # mas o Allan precisa ver -- a segunda linha vence em silêncio.
            colisoes.append((chave, nome))
        chaves_vistas.add(chave)

        setor = _texto(row[idx[COL_SETOR]])
        sub = _texto(row[idx[COL_SUB]])
        grupo = _texto(row[idx[COL_GRUPO]])
        if not setor:
            sem_taxonomia += 1

        if dry_run:
            existente = db.scalar(select(Issuer).where(Issuer.key == chave))
            if existente is None:
                criados += 1
            else:
                atualizados += 1
            continue

        issuer = db.scalar(select(Issuer).where(Issuer.key == chave))
        if issuer is None:
            issuer = Issuer(key=chave, nome=nome, nome_anbima=nome)
            db.add(issuer)
            db.flush()
            criados += 1
        else:
            atualizados += 1
        isv.aplicar_taxonomia(
            issuer,
            setor=setor,
            sub_setor=sub,
            grupo_economico=grupo,
            origem=isv.ORIGEM_SNAPSHOT,
        )

    # Mapa ticker -> nome do emissor, da aba "Tickers". Usado por
    # `vincular_debentures` como caminho de casamento EXATO (ver docstring
    # de lá) -- resolve os casos de grafia que a normalização não pega,
    # tipo "DE SP" vs "DE SÃO PAULO".
    ticker_para_nome: dict[str, str] = {}
    if "Tickers" in wb.sheetnames:
        wt = wb["Tickers"]
        lt = wt.iter_rows(values_only=True)
        ht = [str(c).strip() if c is not None else "" for c in next(lt)]
        if "Ticker" in ht and COL_NOME in ht:
            i_tk, i_nm = ht.index("Ticker"), ht.index(COL_NOME)
            for row in lt:
                tk, nm = _texto(row[i_tk]), _texto(row[i_nm])
                if tk and nm:
                    ticker_para_nome[tk.upper()] = nm

    if not dry_run:
        db.commit()
    return {
        "criados": criados,
        "atualizados": atualizados,
        "sem_taxonomia": sem_taxonomia,
        "ignorados": ignorados,
        "colisoes": colisoes,
        "ticker_para_nome": ticker_para_nome,
    }


# ---------------------------------------------------------------------------
# Ratings
# ---------------------------------------------------------------------------

def carregar_ratings(db, caminho: Path, *, dry_run: bool) -> dict:
    """Lê ações de rating de um .jsonl (checkpoint do scraper) ou .json.

    Cada registro precisa de um nome de emissor (`nome_anbima` ou `nome`)
    e uma data (`dataRating`/`data_acao`); os ratings vêm nos campos
    `fitch`/`sp`/`moodys`, na grafia original da agência.

    Emissor que não casa com nenhum `Issuer` é CONTADO e reportado, mas
    NÃO criado: rating de emissor sem papel nosso é ruído, e criar o
    registro esconderia um erro de casamento (ver docstring de
    `issuers.obter_ou_criar_issuer`).
    """
    registros: list[dict] = []
    texto = caminho.read_text(encoding="utf-8", errors="replace")
    if caminho.suffix == ".jsonl":
        for linha in texto.splitlines():
            linha = linha.strip()
            if linha:
                try:
                    registros.append(json.loads(linha))
                except json.JSONDecodeError:
                    continue
    else:
        dados = json.loads(texto)
        registros = dados if isinstance(dados, list) else dados.get("ratings", [])

    gravadas = duplicadas = sem_emissor = sem_data = 0
    nomes_sem_emissor: dict[str, int] = {}

    for reg in registros:
        nome = _texto(reg.get("nome_anbima")) or _texto(reg.get("nomeEmpresa")) or _texto(reg.get("nome"))
        data_acao = _data(reg.get("dataRating") or reg.get("data_acao") or reg.get("data"))
        if not nome:
            continue
        if data_acao is None:
            sem_data += 1
            continue
        issuer = isv.resolver_issuer(db, nome)
        if issuer is None:
            sem_emissor += 1
            nomes_sem_emissor[nome] = nomes_sem_emissor.get(nome, 0) + 1
            continue
        for campo, agencia in AGENCIA_POR_CAMPO.items():
            rating = _texto(reg.get(campo))
            if not rating:
                continue
            if dry_run:
                gravadas += 1
                continue
            criada = isv.registrar_acao_rating(
                db,
                issuer,
                agencia=agencia,
                rating=rating,
                data_acao=data_acao,
                perspectiva=_texto(reg.get("perspectivaAtual")),
                rating_anterior=_texto(reg.get("ratingAnterior")),
                perspectiva_anterior=_texto(reg.get("perspectivaAnterior")),
                acao=_texto(reg.get("acaoRating")),
                link=_texto(reg.get("link")),
                origem=_texto(reg.get("origem")) or "SNAPSHOT",
            )
            if criada is None:
                duplicadas += 1
            else:
                gravadas += 1
    if not dry_run:
        db.commit()
    return {
        "registros_lidos": len(registros),
        "acoes_gravadas": gravadas,
        "duplicadas": duplicadas,
        "sem_emissor": sem_emissor,
        "sem_data": sem_data,
        "nomes_sem_emissor": nomes_sem_emissor,
    }


# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--taxonomia", type=Path, help="Taxonomia_Emissores.xlsx")
    ap.add_argument("--ratings", type=Path, help=".jsonl/.json com ações de rating")
    ap.add_argument("--dry-run", action="store_true", help="não grava nada")
    ap.add_argument("--apenas-recalcular", action="store_true",
                    help="só recalcula issuer_rating_atual")
    args = ap.parse_args()

    if not any([args.taxonomia, args.ratings, args.apenas_recalcular]):
        ap.error("informe --taxonomia, --ratings ou --apenas-recalcular")

    Base.metadata.create_all(engine)
    run_migrations()

    db = SessionLocal()
    ticker_para_nome: dict[str, str] = {}
    try:
        if args.taxonomia:
            print(f"\n=== TAXONOMIA — {args.taxonomia.name} ===")
            r = carregar_taxonomia(db, args.taxonomia, dry_run=args.dry_run)
            ticker_para_nome = r["ticker_para_nome"]
            print(f"  emissores criados     : {r['criados']}")
            print(f"  emissores atualizados : {r['atualizados']}")
            print(f"  sem setor preenchido  : {r['sem_taxonomia']}")
            print(f"  linhas ignoradas      : {r['ignorados']}")
            if r["colisoes"]:
                print(f"  ATENÇÃO — {len(r['colisoes'])} linhas colapsaram numa chave já vista:")
                for chave, nome in r["colisoes"][:10]:
                    print(f"      [{chave}] <- {nome}")

        if args.ratings:
            print(f"\n=== RATINGS — {args.ratings.name} ===")
            r = carregar_ratings(db, args.ratings, dry_run=args.dry_run)
            print(f"  registros lidos       : {r['registros_lidos']}")
            print(f"  ações gravadas        : {r['acoes_gravadas']}")
            print(f"  já existentes         : {r['duplicadas']}")
            print(f"  sem data válida       : {r['sem_data']}")
            print(f"  emissor não encontrado: {r['sem_emissor']}")
            top = sorted(r["nomes_sem_emissor"].items(), key=lambda x: -x[1])[:10]
            for nome, q in top:
                print(f"      {q:5d}x  {nome}")

        if not args.dry_run:
            print("\n=== VÍNCULO DEBÊNTURE -> EMISSOR ===")
            v = isv.vincular_debentures(db, ticker_para_nome)
            print(f"  debêntures ligadas    : {v['ligadas']}"
                  f"  (por ticker {v['por_ticker']}, por nome {v['por_nome']})")
            if v["aliases_criados"]:
                print(f"  aliases criados       : {len(v['aliases_criados'])}"
                      f"  (grafia do banco != planilha, provado pelo ticker)")
                for nome_banco, nome_issuer in v["aliases_criados"][:8]:
                    print(f"      {nome_banco[:45]:<47} -> {nome_issuer[:40]}")
            if v["revisar"]:
                print(f"  A REVISAR             : {len(v['revisar'])}"
                      f"  (papel ligado pelo ticker, mas os nomes não se parecem —")
                print(f"                          NÃO viraram alias; confirme na Administração)")
                for codigo, nome_banco, nome_issuer in v["revisar"][:10]:
                    print(f"      {codigo:<8} {nome_banco[:38]:<40} -> {nome_issuer[:36]}")
            print(f"  emissores sem match   : {len(v['sem_match'])}")
            for nome, q in sorted(v["sem_match"].items(), key=lambda x: -x[1])[:10]:
                print(f"      {q:3d}x  {nome}")

            print("\n=== RATING VIGENTE ===")
            s = isv.recalcular_todos_ratings(db)
            print(f"  emissores             : {s['emissores']}")
            print(f"  com rating            : {s['com_rating']}")
            ordem = sorted(s["distribuicao"].items(), key=lambda x: -x[1])
            print("  distribuição do rating médio:")
            for rt, q in ordem[:12]:
                print(f"      {rt:<6} {q}")
            if s["desconhecidos"]:
                print("  ratings em formato não reconhecido (ver Administração):")
                for k, q in sorted(s["desconhecidos"].items(), key=lambda x: -x[1])[:10]:
                    print(f"      {q:4d}x  {k}")
        else:
            print("\n[DRY-RUN] nada foi gravado.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
