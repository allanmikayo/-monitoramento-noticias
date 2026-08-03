"""Mapeamento completo de ações de rating local Brasil — Moody's Local,
S&P Global Ratings Brasil e Fitch Ratings — para um período (default:
01/01/2026 até hoje), com detalhe completo por ação (emissor, setor, tipo de
ação, rating anterior/atual, perspectiva anterior/atual).

Por que este script existe (22/07/2026): pedido do Allan que NÃO tem ligação
com o dashboard de monitoramento (esse mapeia só as ~96 empresas cobertas) —
aqui é o MERCADO INTEIRO, todas as ações de rating local das 3 agências no
período. Reaproveita a infraestrutura já construída em app/sources/ (mesmos
sites, mesma técnica Playwright) e data/Setores.xlsx (pra cruzar setor de
quem já é coberto), mas é um script standalone -- não grava no banco do
dashboard, gera direto um .xlsx.

IMPORTANTE — leia antes de rodar/reportar bug: eu (Claude) não tenho acesso a
internet real para estes 3 sites no ambiente onde este código foi escrito
(sandbox bloqueado) -- toda a navegação/estrutura foi confirmada AO VIVO por
mim numa sessão de navegador (Chrome), mas a extração de DETALHE de cada
ação (rating anterior, perspectiva anterior) depende de abrir um PDF por
ação, e eu só confirmei a ESTRUTURA de acesso ao PDF (não testei o parser de
texto do PDF contra centenas de casos reais). A primeira execução de verdade
neste computador é o teste real do parser de detalhe -- se vier muita coisa
"N/D" nas colunas de rating/perspectiva anterior, me manda um trecho do
"texto_fonte" da linha problemática (fica sempre salvo na planilha) que eu
ajusto o regex.

Uso:
    python -m scripts.mapear_ratings_2026                     # 2026-01-01 até hoje, as 3 agências
    python -m scripts.mapear_ratings_2026 --agencia sp         # só uma agência
    python -m scripts.mapear_ratings_2026 --inicio 2026-01-01 --fim 2026-07-22
    python -m scripts.mapear_ratings_2026 --limite 10           # só as 10 primeiras de cada agência (teste rápido)
    python -m scripts.mapear_ratings_2026 --so-listagem          # só lista (rápido), sem abrir PDF/artigo pra detalhe

Saída: data/mapeamento_ratings_<inicio>_<fim>.xlsx
Checkpoint incremental: data/mapeamento_ratings_checkpoint.jsonl (uma linha
JSON por ação já processada -- se o script cair no meio, rodar de novo pula
o que já foi feito em vez de recomeçar do zero).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, date, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.sources.base import USER_AGENT, brt_to_utc  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
SETORES_XLSX = DATA_DIR / "Setores.xlsx"
LOG_PATH = DATA_DIR / "mapear_ratings_2026_log.txt"

MESES_EN = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _log(msg: str) -> None:
    """Imprime na tela E grava em data/mapear_ratings_2026_log.txt (uma
    linha por mensagem, com timestamp) -- criado em 23/07/2026 depois que o
    Allan rodou o .bat e não sabia dizer se tinha funcionado (a janela do
    console rola e não fica nada gravado se ele fechar ou não conseguir
    copiar o texto). Com o log em arquivo, dá pra eu ler
    `data/mapear_ratings_2026_log.txt` direto e diagnosticar exatamente o
    que aconteceu numa rodada, sem depender de descrição de tela. Usado no
    lugar de `print()` em todo o script a partir daqui."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(msg)
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:  # noqa: BLE001 -- log nao pode derrubar a coleta
        pass


# ---------------------------------------------------------------------------
# Modelo de linha de saída
# ---------------------------------------------------------------------------

@dataclass
class AcaoRating:
    agencia: str
    data: str  # ISO yyyy-mm-dd
    emissor: str
    setor: str
    nivel_acao: str  # "Emissor" (rating de risco de crédito do emissor) | "Instrumento" (debênture, FIDC, CRI, CRA etc.)
    tipo_acao: str  # elevação | rebaixamento | afirmação | atribuição | retirada | perspectiva | multiplo/outro
    rating_anterior: str
    rating_atual: str
    perspectiva_anterior: str
    perspectiva_atual: str
    titulo: str
    link: str
    texto_fonte: str = ""  # trecho usado pra extrair rating anterior/perspectiva -- útil pra conferir/corrigir
    id_fonte: str = ""  # id único na fonte (pra dedupe/checkpoint)
    revisar_manualmente: str = ""  # "Sim"/"Não" -- ver `avaliar_qualidade()`
    motivo_revisao: str = ""  # por que ficou marcada pra revisão (vazio se revisar_manualmente="Não")


# ---------------------------------------------------------------------------
# Classificação de tipo de ação e extração via regex (texto do título/corpo)
# ---------------------------------------------------------------------------

_VERBOS_ACAO = [
    (re.compile(r"elev[ao]|upgrade", re.I), "elevação"),
    (re.compile(r"rebaix|downgrade", re.I), "rebaixamento"),
    (re.compile(r"reafirm|afirma|confirma|mant[eé]m", re.I), "afirmação"),
    (re.compile(r"atribu[ií]", re.I), "atribuição (nova)"),
    (re.compile(r"retir", re.I), "retirada"),
    (re.compile(r"perspectiva.*(alterada|revisada)", re.I), "mudança de perspectiva"),
    (re.compile(r"creditwatch|observa[cç][aã]o", re.I), "colocação/remoção em observação"),
    (re.compile(r"diversas a[cç][oõ]es", re.I), "múltiplas ações (conferir manualmente)"),
]

_PERSPECTIVA_RE = re.compile(
    r"perspectiva[^.;]*?\b(est[aá]vel|negativa|positiva|em desenvolvimento)\b", re.I
)


def classificar_acao(titulo: str) -> str:
    for pat, nome in _VERBOS_ACAO:
        if pat.search(titulo):
            return nome
    return "outro (conferir)"


# ---------------------------------------------------------------------------
# Nível da ação (Emissor vs. Instrumento) e extração de emissor a partir do
# TÍTULO -- pedido do Allan em 23/07/2026: "preciso que exista uma coluna que
# me mostre se é uma ação de rating no emissor ou no instrumento (debenture,
# FIDC, etc)" + "preciso que tenha o nome ... do emissor preenchido também".
#
# Por que a partir do TÍTULO e não do texto_fonte (corpo do PDF/artigo):
# testado ao vivo contra os 779 registros já coletados (22/07/2026) -- título
# está sempre presente e é limpo/estruturado nas 3 agências, enquanto
# texto_fonte está 0% preenchido pra S&P e Moody's (bug de extração de PDF,
# ver sp_extract_detail/moodys_extract_detail) e só ~50% "de verdade" pra
# Fitch (o resto é lixo de menu de navegação -- ver fitch_extract_detail). A
# extração por título é mais confiável e funciona igual pras 3 agências.
# ---------------------------------------------------------------------------

_KEYWORDS_INSTRUMENTO_RE = re.compile(
    r"\bFIDC\b|\bCRI\b|\bCRIs\b|\bCRA\b|\bCRAs\b|\bCCB\b|\bCPR-?Fs?\b|"
    r"deb[eê]ntures?|c[eé]dulas?|cotas?\b|s[ée]rie[s]?\b|emiss[ãa]o|emiss[oõ]es|"
    r"certificados de receb[ií]veis|fundo de investimento",
    re.I,
)


def classificar_nivel_acao(titulo: str) -> str:
    """'Instrumento' quando o título menciona claramente um papel/dívida
    específica (debênture, FIDC, CRI, CRA, cotas, série, emissão etc.) --
    nesses casos o rating é da dívida, não necessariamente igual ao rating de
    crédito do emissor/garantidor. Senão, 'Emissor' (rating de risco de
    crédito direto da empresa/entidade)."""
    return "Instrumento" if _KEYWORDS_INSTRUMENTO_RE.search(titulo or "") else "Emissor"


_EMISSOR_STOPWORDS_INICIO = (
    "creditwatch", "emissão", "emissao", "série", "serie", "cotas", "cota",
    "proposta", "grupo", "observação", "observacao", "rating", "ratings",
    "perspectiva", "debênture", "debenture", "debêntures", "debentures",
    "sua", "suas", "seu", "seus", "um", "uma", "diversas", "ações", "acoes",
    "qualidade", "fundo", "texto", "correção", "correcao",
)
_EMISSOR_LEGAL_SUFFIX_RE = re.compile(r"\b(S\.A\.?|S/A|Ltda\.?|FIDC|FII|FIC)\b", re.I)
_EMISSOR_RISCO_PAREN_RE = re.compile(r"\(Risco\s+([^)]+)\)", re.I)
_EMISSOR_RISCO_RE = re.compile(r"\bRisco\s+([A-ZÀ-Ú][^;,.)]*?)(?=\s+para\b|[;,.)]|$)", re.I)
_EMISSOR_CONECTOR_RE = re.compile(
    r"\b(?:da|do|de|à|às|aos)\s+([A-ZÀ-Ú][^;,.]*?)"
    r"(?=\s+(?:da|do|de|à|às|aos)\b|\s+para\b|\s+em\s|\s+lastreadas?\b|[;,]|\.$|$)",
    re.I,
)
_EMISSOR_TRUNCA_RE = re.compile(
    r"\s+(?:"
    r"elevad[oa]s?|rebaixad[oa]s?|reafirmad[oa]s?|confirmad[oa]s?|mantid[oa]s?|"
    r"retirad[oa]s?|removid[oa]s?|atribu[ií]d[oa]s?|segu(?:e|em)\b.*|"
    r"coloca(?:d[oa]s?)?\b.*|subsidi[aá]rias?\b.*|"
    r"e\s+(?:de|da|do|à|às|em|seus?|suas?|subsequentemente|o|a|os|as)\b.*"
    r").*$",
    re.I,
)


def _emissor_e_stopword(candidato: str) -> bool:
    c = candidato.strip().lower()
    return any(c.startswith(sw) for sw in _EMISSOR_STOPWORDS_INICIO) or bool(re.match(r"^\d", c))


def _emissor_limpar(candidato: str) -> str:
    candidato = _EMISSOR_TRUNCA_RE.sub("", candidato)
    candidato = re.sub(r"\s+e$", "", candidato.strip())
    return candidato.strip().rstrip(".,;'\"’‘")


def extrair_emissor_titulo(titulo: str) -> str:
    """Extrai o nome do emissor/entidade de risco de crédito a partir do
    TÍTULO da ação de rating (funciona igual pras 3 agências -- ver
    justificativa acima). Prioridade:
      1. '(Risco X)' ou 'Risco X' -- anotação que a própria agência usa em
         operações estruturadas (CRI/CRA/FIDC) pra indicar a entidade que
         realmente carrega o risco de crédito por trás do papel (ex.: "353ª
         Emissão de CRAs da Eco; Risco Amaggi" -> o risco de crédito real é
         da Amaggi, não da securitizadora Eco que só é a emissora formal).
      2. Padrão 'da/do/de/à/às/aos <Nome>' -- pega candidatos separados por
         essas preposições e FUNDE candidatos adjacentes quando o anterior
         não parece ter terminado (evita perder prefixo em nomes com
         preposição interna, ex.: "Banco do Brasil", "Chapada do Piauí"),
         descarta candidatos que começam com palavra genérica (stopword) e
         prefere o que tem sufixo de razão social (S.A., Ltda, FIDC...);
         sem isso, fica com o último candidato válido (heurística -- nomes
         de empresa costumam vir por último no título das 3 agências).
    Validado manualmente contra amostra aleatória de ~120 títulos reais
    (23/07/2026) -- taxa de "não consegui extrair" ficou em ~4-16% a
    depender da agência (bem menor que os ~15-90% em branco que a extração
    antiga via corpo do PDF/artigo dava)."""
    titulo = titulo or ""
    m = _EMISSOR_RISCO_PAREN_RE.search(titulo)
    if m:
        return _emissor_limpar(m.group(1))
    m = _EMISSOR_RISCO_RE.search(titulo)
    if m:
        return _emissor_limpar(m.group(1))

    matches = list(_EMISSOR_CONECTOR_RE.finditer(titulo))
    fundidos: list[tuple[int, int, str]] = []  # (start, end, texto)
    for m in matches:
        cand = m.group(1)
        if fundidos:
            prev_start, prev_end, prev_text = fundidos[-1]
            gap = titulo[prev_end:m.start()]
            prev_last_word = prev_text.strip().split()[-1] if prev_text.strip() else ""
            if gap.strip() == "" and prev_last_word[:1].isupper() and not _emissor_e_stopword(prev_text):
                connector_word = titulo[m.start():m.start(1)].strip()
                fundidos[-1] = (prev_start, m.end(1), f"{prev_text} {connector_word} {cand}")
                continue
        fundidos.append((m.start(), m.end(1), cand))

    candidatos = []
    for _, _, cand in fundidos:
        cand = _emissor_limpar(cand)
        if cand and not _emissor_e_stopword(cand):
            candidatos.append(cand)
    if not candidatos:
        return ""
    com_sufixo = [c for c in candidatos if _EMISSOR_LEGAL_SUFFIX_RE.search(c)]
    if com_sufixo:
        return com_sufixo[-1]
    return candidatos[-1]


def extrair_perspectiva_atual(texto: str) -> str:
    m = _PERSPECTIVA_RE.search(texto or "")
    return m.group(1).capitalize() if m else ""


_RATING_PATTERNS_CASE_SENSITIVE = [
    re.compile(r"\bbr[A-D][A-Za-z+-]{0,3}(?:\s*\(sf\))?(?:/br[A-Za-z0-9+-]+)?\b"),  # S&P: brAAA, brA-1+, brBB (sf)
]
_RATING_PATTERNS_CASE_INSENSITIVE = [
    re.compile(r"\b[A-C][A-C+-]{0,2}\.br(?:\s*\(sf\))?\b", re.I),  # Moody's: AAA.br, B-.br
    re.compile(r"\b[A-F][A-F0-9+\-]{0,4}(?:sf)?\(bra\)", re.I),  # Fitch: AA(bra), F1+(bra), AAAsf(bra)
]


def extrair_ratings_mencionados(texto: str) -> list[str]:
    """Pega tokens de rating no padrão de escala nacional/local das 3
    agências: S&P 'brAAA', 'brA-1+', Moody's 'AAA.br', 'B-.br', Fitch
    'AA(bra)', 'F1+(bra)'. Não distingue anterior/atual sozinho -- isso é
    inferido depois por posição no texto (ver `parse_texto_para_detalhe`).

    IMPORTANTE: o padrão da S&P (prefixo "br" minúsculo + letra MAIÚSCULA)
    precisa ser lido sem IGNORECASE -- com IGNORECASE ligado, "br[A-D]"
    também bate com o sufixo "(bra)" da Fitch (ex.: o "bra" dentro de
    "AA(bra)"), contaminando o resultado com lixo. Bug encontrado e
    corrigido em 22/07/2026 testando contra títulos reais das 3 agências."""
    texto = texto or ""
    out: list[str] = []
    for p in _RATING_PATTERNS_CASE_SENSITIVE:
        out.extend(p.findall(texto))
    for p in _RATING_PATTERNS_CASE_INSENSITIVE:
        out.extend(p.findall(texto))
    return out


# ---------------------------------------------------------------------------
# Setores.xlsx — cruzamento de setor pra emissores já cobertos pelo dashboard
# ---------------------------------------------------------------------------

def carregar_setores() -> dict[str, str]:
    """Lê data/Setores.xlsx -> {nome_normalizado: setor}. Usado só como
    ATALHO pros ~96 emissores já cobertos pelo dashboard -- pro resto do
    mercado (a maioria das ações aqui, já que este script NÃO filtra por
    cobertura) o setor fica em branco/"" quando a fonte não informa
    (Fitch informa; S&P e Moody's normalmente não, PDF precisaria de
    leitura de contexto que não fazemos aqui)."""
    import openpyxl

    if not SETORES_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(SETORES_XLSX, read_only=True, data_only=True)
    ws = wb.active
    out: dict[str, str] = {}
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        i_setor = header.index("Setor")
        i_comp = header.index("Companhia")
    except ValueError:
        return {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        setor, comp = row[i_setor], row[i_comp]
        if comp:
            out[_norm(str(comp))] = str(setor or "")
    return out


def _norm(s: str) -> str:
    import unicodedata

    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.lower().strip()


def buscar_setor(emissor: str, setores: dict[str, str], setor_direto: str = "") -> str:
    if setor_direto:
        return setor_direto
    if not emissor:
        return ""
    en = _norm(emissor)
    for nome, setor in setores.items():
        if nome and (nome in en or en in nome):
            return setor
    # tenta por primeira palavra significativa (ex.: "Vale do Tijuco" -> "Vale")
    primeira = en.split()[0] if en.split() else ""
    for nome, setor in setores.items():
        if primeira and nome.startswith(primeira):
            return setor
    return ""


# ---------------------------------------------------------------------------
# Checkpoint incremental
# ---------------------------------------------------------------------------

def checkpoint_path() -> Path:
    return DATA_DIR / "mapeamento_ratings_checkpoint.jsonl"


def carregar_checkpoint() -> dict[str, dict]:
    path = checkpoint_path()
    out: dict[str, dict] = {}
    if not path.exists():
        return out
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
                out[row["id_fonte"]] = row
            except (json.JSONDecodeError, KeyError):
                continue
    return out


def salvar_checkpoint_linha(row: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with checkpoint_path().open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")


_MOTIVOS_TIPO_ACAO_AMBIGUO = ("outro (conferir)", "múltiplas ações (conferir manualmente)")


def avaliar_qualidade(row: dict) -> tuple[str, str]:
    """Marca a linha pra revisão manual quando algum campo importante ficou
    fraco/ausente -- pedido implícito do Allan (23/07/2026, "preciso usar
    isso num relatório sell side de credit research"): pra um relatório
    profissional não basta a base estar "quase toda" preenchida, precisa
    dar pra saber EM QUAIS linhas confiar de olho fechado e quais exigem
    conferência manual rápida antes de citar num relatório. Devolve
    (revisar_manualmente, motivo_revisao) -- motivo vem vazio quando não
    precisa revisar."""
    motivos: list[str] = []
    if not (row.get("emissor") or "").strip():
        motivos.append("emissor não identificado")
    if not (row.get("setor") or "").strip():
        motivos.append("setor não identificado")
    ra = (row.get("rating_anterior") or "")
    rt = (row.get("rating_atual") or "")
    if "N/D" in ra or "erro" in ra.lower():
        motivos.append("rating anterior não identificado com confiança")
    if "N/D" in rt or "erro" in rt.lower():
        motivos.append("rating atual não identificado com confiança")
    pa = (row.get("perspectiva_anterior") or "")
    if "N/D" in pa:
        motivos.append("perspectiva anterior não identificada")
    if (row.get("tipo_acao") or "") in _MOTIVOS_TIPO_ACAO_AMBIGUO:
        motivos.append("tipo de ação ambíguo (conferir título)")
    if not (row.get("texto_fonte") or "").strip():
        motivos.append("sem texto fonte pra conferir (extração baseada só no título)")
    revisar = "Sim" if motivos else "Não"
    return revisar, "; ".join(motivos)


def _upgradar_linha_checkpoint(row: dict, setores: dict[str, str]) -> AcaoRating:
    """Reconstrói uma linha a partir do checkpoint em disco, preenchendo de
    graça (sem nenhuma chamada de rede, só regex sobre o título já salvo)
    campos que ficaram em branco/ausentes em rodadas anteriores -- em
    particular:
      - 'nivel_acao': coluna nova (23/07/2026, pedido do Allan) que não
        existe nos 779 registros já coletados antes dela existir.
      - 'emissor': se ficou vazio (bug antigo do extrator baseado no corpo
        do PDF/artigo, corrigido em 23/07/2026 pro extrator baseado no
        título), tenta de novo com `extrair_emissor_titulo`.
      - 'setor': se ficou vazio mas agora dá pra achar emissor, tenta cruzar
        com Setores.xlsx.
    Isso faz uma rodada normal (sem --reprocessar) já vir com essas colunas
    melhoradas pros registros antigos, sem precisar reprocessar nada."""
    row = dict(row)
    titulo = row.get("titulo", "")
    if not row.get("nivel_acao"):
        row["nivel_acao"] = classificar_nivel_acao(titulo)
    if not row.get("emissor"):
        row["emissor"] = extrair_emissor_titulo(titulo)
    if not row.get("setor") and row.get("emissor"):
        row["setor"] = buscar_setor(row["emissor"], setores)
    row["revisar_manualmente"], row["motivo_revisao"] = avaliar_qualidade(row)
    return AcaoRating(**row)


# ---------------------------------------------------------------------------
# Parser de texto extraído de PDF/artigo -> campos estruturados
# ---------------------------------------------------------------------------

_DE_PARA_RE = re.compile(
    r"(?:de|from)\s+['‘]?([A-Za-z0-9+\-\.\(\) ]{1,15}?)['’]?\s+(?:para|to)\s+['‘]?([A-Za-z0-9+\-\.\(\) ]{1,15}?)['’]?[\s,;.]",
    re.I,
)

# Fitch, na prática (confirmado em texto_fonte real, 22/07/2026), costuma
# escrever na ordem INVERSA -- "elevou hoje, para 'AAsf(bra)', de
# 'AA-sf(bra)', o Rating..." (primeiro o rating NOVO, depois o antigo) --
# em vez do "de X para Y" que _DE_PARA_RE cobre. Sem este segundo padrão, o
# rating_anterior/atual da Fitch ficava errado ou caindo pro fallback mais
# fraco (contagem de tokens). Grupo 1 = rating ATUAL (vem depois de
# "para"), grupo 2 = rating ANTERIOR (vem depois de "de").
_PARA_DE_RE = re.compile(
    r"(?:para|to)\s+['‘]?([A-Za-z0-9+\-\.\(\) ]{1,15}?)['’]?\s*,?\s*(?:de|from)\s+['‘]?([A-Za-z0-9+\-\.\(\) ]{1,15}?)['’]?[\s,;.]",
    re.I,
)


def _parece_rating(token: str) -> bool:
    """Valida se um texto capturado pelo regex genérico 'de X para Y'
    realmente parece um código de rating (e não uma frase qualquer tipo
    'de alavancagem para financiar', que o regex genérico também casa por
    engano -- bug encontrado testando contra títulos reais, 22/07/2026).
    Exige que o token bata em algum dos padrões específicos de rating das
    3 agências."""
    token = token.strip()
    if not token:
        return False
    return bool(extrair_ratings_mencionados(token))


def parse_texto_para_detalhe(texto: str, titulo: str) -> dict:
    """Tenta achar rating anterior/atual e perspectiva anterior/atual dentro
    do texto completo (título + corpo do PDF/artigo, quando disponível).
    Estratégia, em ordem:
      1. Padrão explícito "de X para Y" / "from X to Y" (comum em ações de
         elevação/rebaixamento) -> rating_anterior=X, rating_atual=Y --
         só aceito se os dois lados baterem com um padrão de rating de
         verdade (ver `_parece_rating`).
      2. Se não achou o "de/para", pega todos os tokens de rating do texto:
         se só tem 1 valor distinto -> rating_atual=esse valor,
         rating_anterior = mesmo valor (afirmação/atribuição nova).
         Se tem 2+ valores distintos e a ação é elevação/rebaixamento,
         assume o PRIMEIRO como anterior e o ÚLTIMO como atual (heurística
         -- textos de rating costumam citar o histórico em ordem
         cronológica) -- MARCAR pra conferência manual via texto_fonte.
      3. Perspectiva: procura "perspectiva anterior"/"previous outlook" +
         valor; se não achar, assume perspectiva_anterior = perspectiva_atual
         quando ação é afirmação (não muda perspectiva por definição salvo
         dito o contrário no título, ex. "perspectiva revisada")."""
    acao = classificar_acao(titulo)
    perspectiva_atual = extrair_perspectiva_atual(titulo) or extrair_perspectiva_atual(texto)

    rating_anterior = ""
    rating_atual = ""

    m = None
    ordem_invertida = False
    for candidato in _DE_PARA_RE.finditer(texto or ""):
        if _parece_rating(candidato.group(1)) and _parece_rating(candidato.group(2)):
            m = candidato
            break
    if m is None:
        for candidato in _DE_PARA_RE.finditer(titulo or ""):
            if _parece_rating(candidato.group(1)) and _parece_rating(candidato.group(2)):
                m = candidato
                break
    if m is None:
        # tenta a ordem invertida "para X, de Y" (comum na Fitch -- ver
        # docstring de _PARA_DE_RE) antes de cair no fallback mais fraco
        for candidato in _PARA_DE_RE.finditer(texto or ""):
            if _parece_rating(candidato.group(1)) and _parece_rating(candidato.group(2)):
                m = candidato
                ordem_invertida = True
                break
    if m is None:
        for candidato in _PARA_DE_RE.finditer(titulo or ""):
            if _parece_rating(candidato.group(1)) and _parece_rating(candidato.group(2)):
                m = candidato
                ordem_invertida = True
                break
    if m:
        if ordem_invertida:
            rating_atual, rating_anterior = m.group(1).strip(), m.group(2).strip()
        else:
            rating_anterior, rating_atual = m.group(1).strip(), m.group(2).strip()
    else:
        tokens = extrair_ratings_mencionados(titulo) or extrair_ratings_mencionados(texto)
        distintos = list(dict.fromkeys(t.strip() for t in tokens))
        if len(distintos) == 1:
            rating_atual = distintos[0]
            if acao == "afirmação":
                rating_anterior = distintos[0]
            elif acao == "atribuição (nova)":
                rating_anterior = "N/A (rating novo)"
            else:
                rating_anterior = "N/D (só achei 1 valor no texto -- conferir texto_fonte)"
        elif len(distintos) >= 2:
            rating_anterior, rating_atual = distintos[0], distintos[-1]
        else:
            rating_anterior = rating_atual = "N/D"

    if acao == "afirmação":
        perspectiva_anterior = perspectiva_atual
    elif acao == "atribuição (nova)":
        perspectiva_anterior = "N/A (rating novo)"
    else:
        perspectiva_anterior = "N/D (conferir texto_fonte)"

    return {
        "tipo_acao": acao,
        "rating_anterior": rating_anterior,
        "rating_atual": rating_atual,
        "perspectiva_anterior": perspectiva_anterior,
        "perspectiva_atual": perspectiva_atual,
    }


def extrair_pdf_texto(pdf_bytes: bytes) -> str:
    import io

    try:
        import pdfplumber
    except ImportError:
        _log("AVISO: pdfplumber não instalado (pip install pdfplumber) -- não consigo ler PDFs.")
        return ""
    texto_paginas = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for pg in pdf.pages[:3]:  # 3 primeiras páginas costumam bastar (resumo + tabela de ratings)
                t = pg.extract_text() or ""
                texto_paginas.append(t)
    except Exception as e:  # noqa: BLE001
        _log(f"  aviso: falha ao ler PDF ({e})")
    return "\n".join(texto_paginas)


# ---------------------------------------------------------------------------
# S&P Global Ratings Brasil
# ---------------------------------------------------------------------------

SP_URL = "https://brazil.ratings.spglobal.com/ratings/pt/regulatory/press-releases"
SP_API_TEMPLATE = (
    "https://api.use1.prod.ratings.spglobal.com/rbz-nsrbrazilapi/extoauthv2/"
    "brazilRatings/sourceId/{id}?apikey=510153a9-99b2-4028-b1e4-b27d45fde011"
)


def sp_parse_data(texto: str):
    m = re.search(r"(\d{1,2})-([A-Za-z]{3})-(\d{4})", texto or "")
    if not m:
        return None
    dia, mes_abrev, ano = m.groups()
    mes = MESES_EN.get(mes_abrev.lower())
    if not mes:
        return None
    return brt_to_utc(int(ano), mes, int(dia)).date()


def sp_collect_listing(inicio: date, fim: date, limite: int | None = None) -> list[dict]:
    """Reaproveita `app.sources.spglobal.fetch()` (já calibrado e testado
    ao vivo pelo Allan pro dashboard -- inclui o fallback pra Chrome real
    quando o Chromium puro leva bloqueio da Akamai, o filtro 'Últimos 12
    Meses' certo e a paginação). Só ajusta os limites internos do módulo
    (normalmente pensados pra uma varredura de poucos dias) pra cobrir o
    período pedido aqui, que pode ser bem mais longo."""
    from datetime import timedelta as _td

    from app.sources import spglobal

    dias_necessarios = max((date.today() - inicio).days + 10, 30)
    spglobal._MAX_AGE = _td(days=dias_necessarios)
    spglobal._MAX_PAGES = 30  # 25/página -- de sobra pra cobrir 2026 inteiro

    artigos = spglobal.fetch(SP_URL)

    out: list[dict] = []
    for a in artigos:
        if not a.published_at:
            continue
        d = a.published_at.date()
        if not (inicio <= d <= fim):
            continue
        sid = a.url.rstrip("/").split("/")[-1]
        out.append({"id_fonte": sid, "data": d.isoformat(), "titulo": a.title, "link": a.url})
        if limite and len(out) >= limite:
            break
    return out


def sp_extract_detail(item: dict) -> dict:
    """Abre o PDF da ação. Estratégia primária: pedir o PDF direto pela API
    que o próprio site usa (visto ao vivo no devtools -- apikey é público,
    embutido no JS do site). Fallback: abrir a página com Playwright e
    puxar os bytes do PDF renderizado (blob:) via fetch dentro da própria
    página."""
    import requests

    sid = item["id_fonte"]
    pdf_bytes = None
    try:
        resp = requests.get(
            SP_API_TEMPLATE.format(id=sid),
            headers={"User-Agent": USER_AGENT, "Referer": item["link"], "Accept": "application/json, */*"},
            timeout=20,
        )
        if resp.status_code == 200:
            ctype = resp.headers.get("content-type", "")
            if "pdf" in ctype:
                pdf_bytes = resp.content
            else:
                try:
                    payload = resp.json()
                    for key in ("pdfBase64", "content", "data", "documentContent", "fileContent", "base64"):
                        val = payload.get(key) if isinstance(payload, dict) else None
                        if isinstance(val, str) and len(val) > 200:
                            import base64

                            pdf_bytes = base64.b64decode(val)
                            break
                except ValueError:
                    pass
    except Exception as e:  # noqa: BLE001
        _log(f"  aviso: API direta da S&P falhou p/ id={sid} ({e})")

    if pdf_bytes is None:
        pdf_bytes = _sp_fallback_playwright_pdf(item["link"])

    texto = extrair_pdf_texto(pdf_bytes) if pdf_bytes else ""
    detalhe = parse_texto_para_detalhe(texto or item["titulo"], item["titulo"])
    emissor = extrair_emissor_titulo(item["titulo"])
    return {**detalhe, "emissor": emissor, "texto_fonte": (texto or "")[:1200]}


def _sp_fallback_playwright_pdf(url: str) -> bytes | None:
    """Abre a página do artigo (que embute o PDF via object/embed -- o
    dashboard confirma isso ao vivo pro mesmo domínio) e puxa os bytes via
    fetch() dentro da própria página.

    CORREÇÃO (23/07/2026): a versão anterior usava `pw.chromium.launch
    (headless=True)` puro, sem nenhum bypass -- resultado real (checkpoint
    do Allan): 0/139 ações da S&P vieram com texto_fonte, ou seja essa
    função falhava silenciosamente em TODAS as tentativas, não só numa
    instabilidade pontual. `app/sources/spglobal.py` já tinha diagnosticado
    (ao vivo, 17/07/2026) que a S&P bloqueia esse domínio via Akamai Bot
    Manager quando detecta a "impressão digital" do Chromium embutido do
    Playwright -- a correção lá foi tentar primeiro o Chrome de verdade
    instalado na máquina (`channel="chrome"`) antes de cair pro Chromium.
    Esta função nunca tinha recebido esse mesmo fix -- replicado agora,
    junto com o mascaramento de `navigator.webdriver` e o dismiss de banner
    de cookie que também já existiam lá (um banner de cookie bloqueando o
    <object>/<embed> do PDF é outro motivo plausível pra blob_url vir
    None, já que aqui é sempre um browser context NOVO, sem consentimento
    prévio)."""
    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.launch(
                channel="chrome", headless=True, args=["--disable-blink-features=AutomationControlled"]
            )
        except Exception as e:  # noqa: BLE001
            _log(f"  aviso: Chrome real não disponível p/ fallback S&P ({e}) -- usando Chromium embutido")
            browser = pw.chromium.launch(headless=True, args=["--disable-blink-features=AutomationControlled"])
        try:
            ctx = browser.new_context(user_agent=USER_AGENT, viewport={"width": 1366, "height": 900}, locale="pt-BR")
            ctx.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            page = ctx.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(3500)
            for sel in [
                "#onetrust-accept-btn-handler", "button:has-text('Accept All')", "button:has-text('Accept')",
                "button:has-text('Aceitar todos')", "button:has-text('Aceitar')", "#accept-cookies",
                ".cookie-accept", "[data-testid='cookie-accept']",
            ]:
                try:
                    el = page.query_selector(sel)
                    if el and el.is_visible():
                        el.click()
                        page.wait_for_timeout(600)
                        break
                except Exception:  # noqa: BLE001
                    pass
            page.wait_for_timeout(1500)
            blob_url = page.evaluate(
                "document.querySelector('object[type=\"application/pdf\"]')?.data "
                "|| document.querySelector('embed')?.src || null"
            )
            if not blob_url:
                return None
            b64 = page.evaluate(
                """
                async (u) => {
                  const r = await fetch(u);
                  const buf = await r.arrayBuffer();
                  let s = '';
                  const bytes = new Uint8Array(buf);
                  for (let i = 0; i < bytes.length; i++) s += String.fromCharCode(bytes[i]);
                  return btoa(s);
                }
                """,
                blob_url,
            )
            import base64

            return base64.b64decode(b64)
        except Exception as e:  # noqa: BLE001
            _log(f"  aviso: fallback Playwright (blob PDF) falhou ({e})")
            return None
        finally:
            browser.close()


# ---------------------------------------------------------------------------
# Moody's Local Brasil
# ---------------------------------------------------------------------------

MOODYS_URL = "https://moodyslocal.com.br/relatorios/acoes-de-rating/"


def moodys_parse_data(texto: str):
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", texto or "")
    if not m:
        return None
    dia, mes, ano = (int(x) for x in m.groups())
    return brt_to_utc(ano, mes, dia).date()


def moodys_collect_listing(inicio: date, fim: date, limite: int | None = None) -> list[dict]:
    """Reaproveita `app.sources.moodys_local.fetch()` (já calibrado pro
    dashboard -- classes de coluna exatas confirmadas ao vivo, paginação
    por `.paginate_button.next`). Só amplia `_MAX_AGE`/`_MAX_PAGES` pra
    cobrir o período pedido (o valor default do dashboard é uma janela
    curta de poucos dias)."""
    from datetime import timedelta as _td

    from app.sources import moodys_local

    dias_necessarios = max((date.today() - inicio).days + 10, 30)
    moodys_local._MAX_AGE = _td(days=dias_necessarios)
    moodys_local._MAX_PAGES = 60  # ~25/página -- cobre 2026 inteiro com folga

    artigos = moodys_local.fetch(MOODYS_URL)

    out: list[dict] = []
    for a in artigos:
        if not a.published_at:
            continue
        d = a.published_at.date()
        if not (inicio <= d <= fim):
            continue
        sid = a.url.rstrip("/").split("/")[-1] or a.title
        out.append({"id_fonte": sid, "data": d.isoformat(), "titulo": a.title, "link": a.url})
        if limite and len(out) >= limite:
            break
    return out


def moodys_extract_detail(item: dict) -> dict:
    """Abre a página do artigo (HTML simples, só tem título/data/botão
    Download) e baixa o PDF pelo link direto de wp-content/uploads/ pra
    extrair o texto completo (a página em si não tem o corpo do relatório --
    confirmado ao vivo, 22/07/2026).

    CORREÇÃO (23/07/2026): a versão anterior usava `requests.get()` puro
    pras duas chamadas (página do artigo e download do PDF) -- resultado
    real (checkpoint do Allan): 0/262 ações da Moody's vieram com
    texto_fonte, sempre caindo no `except` silenciosamente. O próprio
    `app/sources/base.py` já documenta o motivo mais provável: "Moody's
    Local, atrás de proteção anti-robo, devolve 403 Forbidden para
    requests comuns mesmo com headers de navegador corretos -- o bloqueio
    acontece no handshake TLS" -- por isso o resto do projeto usa
    `curl_cffi` (que imita o TLS fingerprint de um Chrome de verdade) via
    `app.sources.base.get()`, e não `requests` puro. Esta função nunca
    tinha recebido esse mesmo tratamento -- corrigido agora."""
    from bs4 import BeautifulSoup

    from app.sources.base import get as _base_get

    texto = ""
    try:
        resp = _base_get(item["link"])
        soup = BeautifulSoup(resp.content, "lxml")
        pdf_link = None
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.lower().endswith(".pdf") or "wp-content/uploads" in href:
                pdf_link = href
                break
        if pdf_link:
            if pdf_link.startswith("/"):
                pdf_link = "https://moodyslocal.com.br" + pdf_link
            pdf_resp = _base_get(pdf_link)
            texto = extrair_pdf_texto(pdf_resp.content)
    except Exception as e:  # noqa: BLE001
        _log(f"  aviso: falha ao baixar/ler PDF da Moody's p/ {item['link']} ({e})")

    detalhe = parse_texto_para_detalhe(texto or item["titulo"], item["titulo"])
    emissor = extrair_emissor_titulo(item["titulo"])
    return {**detalhe, "emissor": emissor, "texto_fonte": (texto or "")[:1200]}


# ---------------------------------------------------------------------------
# Fitch Ratings (RAC Portuguese / Brazil)
# ---------------------------------------------------------------------------

FITCH_SEARCH_URL = (
    "https://www.fitchratings.com/search?dateValue={datevalue}&expanded=racs"
    "&filter.sector=&filter.language=Portuguese&filter.region=&filter.country="
    "&filter.reportType=Rating+Action+Commentary&filter.topic=&viewType=data"
)


def fitch_parse_data(dia: str, mes_abrev: str, ano: str):
    mes = MESES_EN.get(mes_abrev.lower()[:3])
    if not mes:
        return None
    return brt_to_utc(int(ano), mes, int(dia)).date()


_FITCH_DATE_RE = re.compile(r"(\d{1,2})\s+([A-Za-z]{3})[a-z]*\s+(\d{4})")


def fitch_collect_listing(inicio: date, fim: date, limite: int | None = None) -> list[dict]:
    """Coleta a listagem da Fitch com o PRÓPRIO scraper (em vez de só
    reaproveitar `app.sources.fitch.fetch()`) pra também capturar a coluna
    "Sector & Country", que só aparece no modo `viewType=data` da URL de
    busca e que o `fetch()` do dashboard não lê (ele só pega título/data/
    link, que é tudo que o dashboard precisa).

    MELHORIA (23/07/2026, pedido do Allan pra uso em relatório sell side):
    confirmado AO VIVO, navegando de verdade no site (não é chute), que:
      - `.frw-article-data--title` e `.frw-article-data--sector` sempre
        aparecem em listas PARALELAS na mesma ordem (índice 0 = cabeçalho
        da tabela "TITLE"/"SECTOR & COUNTRY", índice 1..N = uma linha por
        ação, mesma posição nas duas listas).
      - o setor vem em até 3 níveis dentro de um único <p> separados por
        <br> (ex.: "Corporate Finance" / "Healthcare and Pharma" /
        "Brazil" -- setor macro / subsetor / país; op. estruturada costuma
        vir só com 2 níveis, ex. "Structured Finance" / "Structured
        Finance: ABS" / "Brazil").
      - a data vem em 2 <span> (dia, e mês+ano com <br> no meio) que juntam
        pra "22 Jul 2026" usando `get_text(" ", strip=True)` -- mesmo
        padrão que `_DATE_RE` do módulo `app/sources/fitch.py` já espera,
        reaproveitado aqui como `_FITCH_DATE_RE`.
    Isso dá setor de VERDADE (direto da própria classificação da Fitch)
    pra TODA a base da Fitch, sem depender do Setores.xlsx (que só cobre
    as ~96 empresas do dashboard) -- maior fonte de linhas do mapeamento.

    Reaproveita só as funções auxiliares já calibradas de
    `app.sources.fitch` (dismiss de cookie banner e paginação), não o
    `fetch()` inteiro."""
    from playwright.sync_api import sync_playwright
    from bs4 import BeautifulSoup

    from app.sources import fitch as _fitch_mod

    url = FITCH_SEARCH_URL.format(datevalue="lastYear")
    out: list[dict] = []
    seen: set[str] = set()

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.launch(headless=True, args=["--disable-blink-features=AutomationControlled"])
        except Exception as e:  # noqa: BLE001
            raise RuntimeError(f"não consegui abrir o navegador (Chromium) pra Fitch: {e}") from e
        try:
            ctx = browser.new_context(user_agent=USER_AGENT, viewport={"width": 1366, "height": 900}, locale="pt-BR")
            ctx.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            page = ctx.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(3000)
            _fitch_mod._dismiss_cookie_banner(page)
            page.wait_for_timeout(1000)
            try:
                page.wait_for_selector(".frw-article-data--title a[href]", timeout=20000)
            except Exception:
                _log("  aviso: listagem da Fitch não apareceu em 20s -- pode ter mudado o layout do site")

            for pagina in range(1, 81):  # ~10 itens/página -- 80 páginas cobre 2026 inteiro com folga
                html = page.content()
                soup = BeautifulSoup(html, "lxml")
                titulos = soup.select(".frw-article-data--title")
                setores = soup.select(".frw-article-data--sector")
                datas = soup.select(".frw-article-data--date")
                novos_nesta_pagina = 0
                for i in range(1, len(titulos)):  # índice 0 = cabeçalho da tabela, pula
                    tcell = titulos[i]
                    a = tcell.select_one("a[href]")
                    if not a:
                        continue
                    href = a.get("href") or ""
                    if href.startswith("/"):
                        href = "https://www.fitchratings.com" + href
                    if not href or href in seen:
                        continue
                    titulo_txt = (a.get("aria-label") or a.get_text(" ", strip=True) or "").strip()
                    if not titulo_txt:
                        continue

                    data_iso = ""
                    if i < len(datas):
                        m = _FITCH_DATE_RE.search(datas[i].get_text(" ", strip=True))
                        if m:
                            dia, mes_abrev, ano = m.groups()
                            d = fitch_parse_data(dia, mes_abrev, ano)
                            if d:
                                data_iso = d.isoformat()
                                if not (inicio <= d <= fim):
                                    seen.add(href)
                                    continue

                    setor_txt = ""
                    if i < len(setores):
                        setor_txt = setores[i].get_text(" | ", strip=True)

                    seen.add(href)
                    out.append({
                        "id_fonte": href, "data": data_iso, "titulo": titulo_txt, "link": href,
                        "setor_direto": setor_txt,
                    })
                    novos_nesta_pagina += 1
                    if limite and len(out) >= limite:
                        break

                if limite and len(out) >= limite:
                    break
                if len(titulos) <= 1:
                    break  # página vazia (só o cabeçalho) -- acabaram os resultados
                if not _fitch_mod._clicar_proxima_pagina(page):
                    break
        finally:
            browser.close()

    return out


_FITCH_MARCADOR_CORPO = "RATING ACTION COMMENTARY"


def fitch_extract_detail(item: dict) -> dict:
    """Abre o artigo (HTML, não é PDF -- confirmado ao vivo que é
    Rating Action Commentary renderizado em React/HTML normal) e extrai o
    texto completo pra achar rating anterior/atual e perspectiva.

    CORREÇÃO (23/07/2026): analisando o checkpoint real do Allan, só
    188/378 ações da Fitch tinham corpo de verdade -- as outras 190
    "tinham" texto_fonte não-vazio, mas era só o menu de navegação/rodapé
    do site (a página ainda não tinha montado o corpo do artigo quando
    `page.inner_text("body")` rodou, aos 3s de espera). Como não é uma
    exceção, isso ficava SILENCIOSAMENTE salvo no checkpoint como sucesso,
    travando aquela ação pra sempre com dados incompletos mesmo depois de
    aumentar o tempo de espera. Agora: espera mais, tenta de novo uma vez
    se o marcador "RATING ACTION COMMENTARY" (sempre presente no corpo de
    verdade) não aparecer, e se ainda assim não aparecer, LEVANTA exceção
    -- isso aciona o mecanismo de retentativa que já existe em `run()` e
    garante que a ação não fica marcada como concluída até vir com corpo
    de verdade (ver `limpar_cache_sem_texto` pra destravar ações que já
    ficaram presas nesse estado em rodadas anteriores)."""
    from playwright.sync_api import sync_playwright

    texto = ""
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        try:
            ctx = browser.new_context(user_agent=USER_AGENT, locale="pt-BR")
            page = ctx.new_page()
            page.goto(item["link"], wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(3500)
            texto = page.inner_text("body")
            if _FITCH_MARCADOR_CORPO not in texto:
                # corpo ainda não montou -- espera mais um pouco e tenta de novo
                page.wait_for_timeout(4000)
                texto = page.inner_text("body")
        finally:
            browser.close()

    if _FITCH_MARCADOR_CORPO not in texto:
        raise RuntimeError(
            f"corpo do artigo não carregou a tempo (sem '{_FITCH_MARCADOR_CORPO}' no texto capturado)"
        )

    detalhe = parse_texto_para_detalhe(texto or item["titulo"], item["titulo"])
    emissor = extrair_emissor_titulo(item["titulo"])
    return {**detalhe, "emissor": emissor, "texto_fonte": (texto or "")[:1200]}


# ---------------------------------------------------------------------------
# Orquestração
# ---------------------------------------------------------------------------

COLETORES = {
    "sp": ("S&P Global Ratings Brasil", sp_collect_listing, sp_extract_detail),
    "moodys": ("Moody's Local Brasil", moodys_collect_listing, moodys_extract_detail),
    "fitch": ("Fitch Ratings", fitch_collect_listing, fitch_extract_detail),
}

# Retentativas generosas -- pedido explícito do Allan (23/07/2026): "não tem
# problema demorar, se o processo for realmente robusto". Preferimos gastar
# mais tempo tentando de novo (rede instável, bloqueio passageiro de bot) a
# desistir cedo e deixar a linha incompleta numa base pra relatório sell
# side, onde completude/precisão importam mais que velocidade.
_TENTATIVAS_LISTAGEM = 4
_ESPERA_LISTAGEM_S = 25
_TENTATIVAS_ITEM = 4
_ESPERA_ITEM_S = (5, 15, 35)  # backoff crescente entre as tentativas de detalhe


def preflight_checks(agencias: list[str]) -> None:
    """Roda ANTES da coleta longa -- testa em segundos se dá pra alcançar
    cada site e se o navegador (Playwright) abre normalmente, e grava tudo
    no log. Criado em 23/07/2026 depois de uma rodada em que o Allan não
    sabia dizer se tinha funcionado: com isso, um problema de rede/
    instalação aparece logo nas primeiras linhas do log em vez de eu ter
    que adivinhar depois de uma rodada de horas que não gerou nada."""
    _log("=" * 60)
    _log("Checagem inicial (preflight) -- rede e navegador")
    _log("=" * 60)

    from app.sources.base import get as _base_get

    sites = {
        "sp": ("S&P Global Ratings Brasil", SP_URL),
        "moodys": ("Moody's Local Brasil", MOODYS_URL),
        "fitch": ("Fitch Ratings", "https://www.fitchratings.com"),
    }
    for chave in agencias:
        nome, url = sites.get(chave, (chave, None))
        if not url:
            continue
        try:
            resp = _base_get(url)
            _log(f"  rede -> {nome}: OK (status {resp.status_code})")
        except Exception as e:  # noqa: BLE001
            _log(f"  rede -> {nome}: FALHOU ({e}) -- pode ser sua internet ou bloqueio do site; "
                 f"a coleta vai tentar mesmo assim (com retentativas)")

    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as pw:
            try:
                b = pw.chromium.launch(channel="chrome", headless=True)
                b.close()
                _log("  navegador -> Chrome real (usado no fallback da S&P): OK")
            except Exception as e:  # noqa: BLE001
                _log(f"  navegador -> Chrome real: indisponível ({e}) -- vai usar Chromium; se quiser "
                     f"instalar, rode 'python -m playwright install chrome' com o ambiente do projeto ativado")
            try:
                b = pw.chromium.launch(headless=True)
                b.close()
                _log("  navegador -> Chromium: OK")
            except Exception as e:  # noqa: BLE001
                _log(f"  navegador -> Chromium: FALHOU ({e}) -- ISSO PRECISA SER RESOLVIDO antes de "
                     f"continuar (rode 'python -m playwright install chromium' com o ambiente do "
                     f"projeto ativado). A coleta vai tentar mesmo assim, mas provavelmente vai falhar "
                     f"em toda ação que precisa abrir o navegador.")
    except ImportError as e:
        _log(f"  navegador -> Playwright não está instalado ({e}) -- rode 'pip install -r requirements.txt'")

    _log("=" * 60 + "\n")


def _formatar_duracao(segundos: float) -> str:
    segundos = int(segundos)
    if segundos < 60:
        return f"{segundos}s"
    minutos, s = divmod(segundos, 60)
    if minutos < 60:
        return f"{minutos}min{s:02d}s"
    horas, m = divmod(minutos, 60)
    return f"{horas}h{m:02d}min"


def run(agencias: list[str], inicio: date, fim: date, limite: int | None, so_listagem: bool) -> list[AcaoRating]:
    preflight_checks(agencias)

    setores = carregar_setores()
    checkpoint = carregar_checkpoint()
    resultado: list[AcaoRating] = []

    for chave in agencias:
        nome_agencia, coletar, detalhar = COLETORES[chave]
        _log(f"\n=== {nome_agencia}: coletando listagem ({inicio} a {fim}) ===")
        listagem: list[dict] = []
        ultimo_erro: Exception | None = None
        for tentativa in range(1, _TENTATIVAS_LISTAGEM + 1):
            try:
                listagem = coletar(inicio, fim, limite)
                ultimo_erro = None
                break
            except Exception as e:  # noqa: BLE001
                ultimo_erro = e
                _log(f"  aviso: tentativa {tentativa}/{_TENTATIVAS_LISTAGEM} falhou ao coletar "
                     f"listagem de {nome_agencia} ({e})")
                if tentativa < _TENTATIVAS_LISTAGEM:
                    _log(f"  aguardando {_ESPERA_LISTAGEM_S}s antes de tentar de novo "
                         f"(pode ser instabilidade passageira de rede)...")
                    time.sleep(_ESPERA_LISTAGEM_S)
        if ultimo_erro is not None:
            _log(
                f"ERRO ao coletar listagem de {nome_agencia} depois de {_TENTATIVAS_LISTAGEM} tentativas: "
                f"{ultimo_erro}\n"
                f"  O que já tinha sido processado de {nome_agencia} em rodadas anteriores continua "
                f"na planilha final (vem do checkpoint) -- só não dá pra buscar ações NOVAS dessa "
                f"agência nesta rodada. Rode o .bat de novo quando a internet estiver estável."
            )
            continue
        _log(f"  {len(listagem)} ações encontradas no período.")

        ja_em_cache = sum(1 for item in listagem if f"{chave}:{item['id_fonte']}" in checkpoint)
        a_buscar = len(listagem) - ja_em_cache
        _log(f"  {ja_em_cache} já estavam no checkpoint (reaproveitadas) -- {a_buscar} vão ser buscadas agora.")
        inicio_agencia = time.time()
        buscadas_ate_agora = 0

        for i, item in enumerate(listagem, start=1):
            cache_key = f"{chave}:{item['id_fonte']}"
            if cache_key in checkpoint:
                resultado.append(_upgradar_linha_checkpoint(checkpoint[cache_key]["row"], setores))
                continue

            buscadas_ate_agora += 1
            elapsed = time.time() - inicio_agencia
            eta_txt = ""
            if buscadas_ate_agora > 1:
                media = elapsed / buscadas_ate_agora
                restante = media * (a_buscar - buscadas_ate_agora)
                eta_txt = f" | decorrido {_formatar_duracao(elapsed)}, falta ~{_formatar_duracao(restante)}"
            _log(f"  [{buscadas_ate_agora}/{a_buscar}]{eta_txt} {item['titulo'][:70]}")

            if so_listagem:
                detalhe = {
                    "tipo_acao": classificar_acao(item["titulo"]),
                    "rating_anterior": "",
                    "rating_atual": "",
                    "perspectiva_anterior": "",
                    "perspectiva_atual": extrair_perspectiva_atual(item["titulo"]),
                    "emissor": extrair_emissor_titulo(item["titulo"]),
                    "texto_fonte": "",
                }
                falha_extracao = False
            else:
                falha_extracao = False
                ultimo_erro_item: Exception | None = None
                detalhe = None
                for tentativa_item in range(1, _TENTATIVAS_ITEM + 1):
                    try:
                        detalhe = detalhar(item)
                        ultimo_erro_item = None
                        break
                    except Exception as e:  # noqa: BLE001
                        ultimo_erro_item = e
                        if tentativa_item < _TENTATIVAS_ITEM:
                            espera = _ESPERA_ITEM_S[min(tentativa_item - 1, len(_ESPERA_ITEM_S) - 1)]
                            _log(f"    tentativa {tentativa_item}/{_TENTATIVAS_ITEM} falhou ({e}) -- "
                                 f"tentando de novo em {espera}s...")
                            time.sleep(espera)
                if ultimo_erro_item is not None:
                    _log(f"    aviso: falha ao extrair detalhe depois de {_TENTATIVAS_ITEM} tentativas "
                         f"({ultimo_erro_item}) -- salvando só o básico do título")
                    falha_extracao = True
                    detalhe = {
                        "tipo_acao": classificar_acao(item["titulo"]),
                        "rating_anterior": "N/D (erro na extração)",
                        "rating_atual": "N/D (erro na extração)",
                        "perspectiva_anterior": "N/D (erro na extração)",
                        "perspectiva_atual": extrair_perspectiva_atual(item["titulo"]),
                        "emissor": extrair_emissor_titulo(item["titulo"]),
                        "texto_fonte": "",
                    }

            emissor = detalhe.get("emissor", "") or extrair_emissor_titulo(item["titulo"])
            setor = buscar_setor(emissor, setores, item.get("setor_direto", ""))
            row_dict = {
                "agencia": nome_agencia,
                "data": item.get("data", ""),
                "emissor": emissor,
                "setor": setor,
                "nivel_acao": classificar_nivel_acao(item["titulo"]),
                "tipo_acao": detalhe.get("tipo_acao", ""),
                "rating_anterior": detalhe.get("rating_anterior", ""),
                "rating_atual": detalhe.get("rating_atual", ""),
                "perspectiva_anterior": detalhe.get("perspectiva_anterior", ""),
                "perspectiva_atual": detalhe.get("perspectiva_atual", ""),
                "titulo": item["titulo"],
                "link": item["link"],
                "texto_fonte": detalhe.get("texto_fonte", ""),
                "id_fonte": cache_key,
            }
            row_dict["revisar_manualmente"], row_dict["motivo_revisao"] = avaliar_qualidade(row_dict)
            row = AcaoRating(**row_dict)
            resultado.append(row)
            if falha_extracao:
                # NÃO salva no checkpoint de propósito -- se ficasse salvo, uma
                # instabilidade de rede passageira (foi o que aconteceu com o
                # Allan em 22/07/2026 -- internet caiu no meio da S&P) travaria
                # essa ação pra sempre como "N/D (erro na extração)", mesmo
                # rodando o script de novo depois. Sem cache, a próxima rodada
                # tenta essa ação de novo do zero.
                _log(f"    (não guardado no checkpoint -- vai tentar de novo na próxima rodada)")
            else:
                salvar_checkpoint_linha({"id_fonte": cache_key, "row": asdict(row)})
            time.sleep(0.3)  # não martelar os sites

        if buscadas_ate_agora:
            _log(f"  {nome_agencia}: {buscadas_ate_agora} ações buscadas em "
                 f"{_formatar_duracao(time.time() - inicio_agencia)}.")

    return resultado


def gerar_xlsx(linhas: list[AcaoRating], caminho: Path) -> None:
    import pandas as pd

    df = pd.DataFrame([asdict(l) for l in linhas])
    if not df.empty:
        df = df.sort_values(["agencia", "data"], ascending=[True, False])
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        cols = [
            "agencia", "data", "emissor", "setor", "nivel_acao", "tipo_acao",
            "rating_anterior", "rating_atual", "perspectiva_anterior", "perspectiva_atual",
            "revisar_manualmente", "motivo_revisao", "titulo", "link", "texto_fonte",
        ]
        df_out = df[cols] if not df.empty else pd.DataFrame(columns=cols)
        df_out.to_excel(writer, sheet_name="Todas as ações", index=False)

        # Aba dedicada só com as linhas marcadas pra revisão manual (ver
        # `avaliar_qualidade()`) -- pedido implícito do Allan (23/07/2026,
        # uso em relatório sell side de credit research): não basta a base
        # estar preenchida, precisa ser rápido de achar as poucas linhas
        # que exigem conferência antes de citar num relatório, sem precisar
        # vasculhar tudo.
        se_revisar = df_out[df_out["revisar_manualmente"] == "Sim"] if not df_out.empty else df_out
        se_revisar.to_excel(writer, sheet_name="Revisar manualmente", index=False)

        for agencia in df_out["agencia"].unique() if not df_out.empty else []:
            aba = agencia[:31]
            df_out[df_out["agencia"] == agencia].to_excel(writer, sheet_name=aba, index=False)

    if not df.empty:
        n_revisar = int((df["revisar_manualmente"] == "Sim").sum())
        _log(f"  qualidade: {n_revisar}/{len(df)} linhas marcadas p/ revisão manual "
             f"({100 * n_revisar / len(df):.1f}%) -- ver aba 'Revisar manualmente'")
    _log(f"\nPlanilha salva em: {caminho}")


def reprocessar_checkpoint(saida: Path | None) -> None:
    """Reprocessa TODO o checkpoint já salvo em disco -- SEM nenhuma
    chamada de rede -- reaplicando as regras mais atuais (23/07/2026):
    nível da ação (emissor/instrumento), extração de emissor a partir do
    título (bem mais confiável que a antiga, baseada no corpo do PDF/
    artigo), cruzamento de setor, e reparse de rating anterior/atual e
    perspectiva a partir do texto_fonte já salvo (pega o fix do regex
    'de X para Y' que agora também entende a ordem reversa usada pela
    Fitch: 'elevou hoje, para X, de Y').

    Serve pra já deixar os 779 registros coletados antes dessas correções
    mais analisáveis SEM esperar uma nova rodada de coleta (que ainda
    precisa dos fixes de extração de PDF da S&P/Moody's rodando de novo
    com internet estável -- isso sim continua exigindo uma rodada normal
    depois). Sobrescreve o checkpoint em disco com os campos atualizados."""
    setores = carregar_setores()
    path = checkpoint_path()
    if not path.exists():
        _log("Nenhum checkpoint encontrado ainda em data/mapeamento_ratings_checkpoint.jsonl -- rode uma coleta primeiro.")
        return

    with path.open("r", encoding="utf-8") as f:
        entradas = [json.loads(linha) for linha in f if linha.strip()]

    _log(f"Reprocessando {len(entradas)} ações já coletadas (sem acessar a internet)...")
    atualizadas: list[dict] = []
    linhas_obj: list[AcaoRating] = []
    for entrada in entradas:
        row = dict(entrada["row"])
        titulo = row.get("titulo", "")
        texto = row.get("texto_fonte", "")

        detalhe = parse_texto_para_detalhe(texto or titulo, titulo)
        row["tipo_acao"] = detalhe["tipo_acao"]
        row["rating_anterior"] = detalhe["rating_anterior"]
        row["rating_atual"] = detalhe["rating_atual"]
        row["perspectiva_anterior"] = detalhe["perspectiva_anterior"]
        row["perspectiva_atual"] = detalhe["perspectiva_atual"]
        row["nivel_acao"] = classificar_nivel_acao(titulo)

        emissor_novo = extrair_emissor_titulo(titulo)
        if emissor_novo:
            row["emissor"] = emissor_novo
        row["setor"] = buscar_setor(row.get("emissor", ""), setores) or row.get("setor", "")
        row["revisar_manualmente"], row["motivo_revisao"] = avaliar_qualidade(row)

        atualizadas.append({"id_fonte": entrada["id_fonte"], "row": row})
        linhas_obj.append(AcaoRating(**row))

    with path.open("w", encoding="utf-8") as f:
        for entrada in atualizadas:
            f.write(json.dumps(entrada, ensure_ascii=False) + "\n")
    _log(f"Checkpoint atualizado em disco ({len(atualizadas)} linhas).")

    saida = saida or DATA_DIR / f"mapeamento_ratings_reprocessado_{date.today().isoformat()}.xlsx"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    gerar_xlsx(linhas_obj, saida)
    _log(f"\nTotal de ações reprocessadas: {len(linhas_obj)}")


def limpar_cache_sem_texto(agencias: list[str]) -> int:
    """Remove do checkpoint entradas marcadas como 'concluídas' mas sem
    texto_fonte de verdade -- bug encontrado em 23/07/2026: falhas de
    extração de PDF/artigo que não geram exceção (S&P e Moody's sempre
    tinham esse problema; Fitch quando a página não carrega o corpo a
    tempo) ficavam com texto_fonte vazio ou só com lixo de menu de
    navegação, mas MESMO ASSIM eram salvas no checkpoint como sucesso --
    travando aquela ação pra sempre com dados incompletos (rating
    anterior, perspectiva anterior) mesmo depois de corrigido o bug de
    extração (ver sp_extract_detail/moodys_extract_detail/
    fitch_extract_detail). Rode com --refazer-sem-texto ANTES de uma coleta
    pra forçar essas ações específicas a serem buscadas de novo com o
    código corrigido -- as que já têm texto_fonte de verdade (a maioria da
    Fitch) não são mexidas, e o resto do checkpoint (outras agências, se
    não pedidas) fica intacto."""
    path = checkpoint_path()
    if not path.exists():
        return 0
    with path.open("r", encoding="utf-8") as f:
        entradas = [json.loads(linha) for linha in f if linha.strip()]

    mantidas: list[dict] = []
    removidas = 0
    for entrada in entradas:
        chave = entrada["id_fonte"]
        prefixo = chave.split(":", 1)[0]
        if prefixo not in agencias:
            mantidas.append(entrada)
            continue
        texto = entrada["row"].get("texto_fonte", "")
        if prefixo == "fitch":
            valido = _FITCH_MARCADOR_CORPO in texto
        else:  # sp, moodys -- nunca tiveram extração de texto funcionando
            valido = bool(texto.strip())
        if valido:
            mantidas.append(entrada)
        else:
            removidas += 1

    with path.open("w", encoding="utf-8") as f:
        for entrada in mantidas:
            f.write(json.dumps(entrada, ensure_ascii=False) + "\n")
    return removidas


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--agencia", choices=["sp", "moodys", "fitch", "all"], default="all")
    ap.add_argument("--inicio", default="2026-01-01")
    ap.add_argument("--fim", default=date.today().isoformat())
    ap.add_argument("--limite", type=int, default=None, help="limita nº de ações por agência (teste rápido)")
    ap.add_argument("--so-listagem", action="store_true", help="não abre PDF/artigo -- só o que dá pra tirar do título")
    ap.add_argument(
        "--reprocessar", action="store_true",
        help="não coleta nada novo -- só reaplica as regras mais atuais (emissor/setor/nível da ação/ratings) "
             "sobre o checkpoint já salvo em disco e gera a planilha na hora (rápido, sem internet)",
    )
    ap.add_argument(
        "--refazer-sem-texto", action="store_true",
        help="antes de coletar, remove do checkpoint ações que ficaram sem texto_fonte de verdade "
             "(bug de extração de PDF/artigo corrigido em 23/07/2026) pra forçar buscar de novo nesta rodada",
    )
    ap.add_argument("--saida", default=None)
    args = ap.parse_args()

    _log("\n" + "#" * 60)
    _log(f"# NOVA RODADA -- {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} -- args: {vars(args)}")
    _log("#" * 60)

    if args.reprocessar:
        reprocessar_checkpoint(Path(args.saida) if args.saida else None)
        return

    inicio = datetime.strptime(args.inicio, "%Y-%m-%d").date()
    fim = datetime.strptime(args.fim, "%Y-%m-%d").date()
    agencias = list(COLETORES.keys()) if args.agencia == "all" else [args.agencia]

    if args.refazer_sem_texto:
        n = limpar_cache_sem_texto(agencias)
        _log(f"Removidas {n} ações do checkpoint (sem texto_fonte de verdade) -- serão buscadas de novo nesta rodada.")

    linhas = run(agencias, inicio, fim, args.limite, args.so_listagem)

    # Reconstrói a lista final a partir do checkpoint completo, não só do
    # que esta rodada conseguiu buscar -- importante quando a COLETA DA
    # LISTAGEM de uma agência falha inteira (ex.: internet caiu no meio,
    # caso real do Allan em 22/07/2026): sem isso, o resultado dessa
    # agência em rodadas anteriores sumiria da planilha desta rodada,
    # mesmo estando salvo em disco.
    setores = carregar_setores()
    checkpoint_final = carregar_checkpoint()
    ja_incluidos = {l.id_fonte for l in linhas}
    for cache_key, entry in checkpoint_final.items():
        if cache_key not in ja_incluidos:
            linhas.append(_upgradar_linha_checkpoint(entry["row"], setores))

    saida = Path(args.saida) if args.saida else DATA_DIR / f"mapeamento_ratings_{inicio}_{fim}.xlsx"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    gerar_xlsx(linhas, saida)
    _log(f"\nTotal de ações mapeadas: {len(linhas)}")


if __name__ == "__main__":
    main()
